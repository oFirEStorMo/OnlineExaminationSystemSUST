import datetime
import jwt
import pytz
from django.shortcuts import render
from django.db import transaction
from google.auth.transport import requests
from google.oauth2 import id_token
from rest_framework.exceptions import AuthenticationFailed, NotAcceptable, NotFound, ParseError
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework import status
from django.db.models import Q
from django.db.models import Subquery
from django.db.models import Sum
from .models import *
from .serializers import *
import mutagen
import speech_recognition as sr
from django.core.files import File
from random import randint
from django.http import FileResponse
from openpyxl import Workbook
import os
import moviepy.editor as moviepy
import moviepy.video.io.ffmpeg_tools as ffmpeg_tools
import random
from .utils import is_correct_answer
# Create your views here.

CLIENT_ID = '405859311003-h78v02m4e3t2bpbksovtfapl6uc27uof.apps.googleusercontent.com'
SECRET_KEY = '_jwt_auth_oes_'


# Authorize Google User


def authorizeUser(request):
    token = request.COOKIES.get('jwt')

    if not token:
        raise AuthenticationFailed('Unauthenticated')

    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=['HS256'])
        return payload
    except jwt.ExpiredSignatureError:
        raise AuthenticationFailed('Unauthenticated')

# Check for the existence of a Exam Cohort and an authorized Evaluator


def authorizeEvaluator(request):
    payload = authorizeUser(request)
    user = payload['id']
    examCohort = request.data['id']
    if examCohort is None:
        raise NotAcceptable(detail='Invalid ExamCohort')
    if Evaluator.objects.filter(user=user, examCohort=examCohort).exists():
        return payload
    raise AuthenticationFailed('Unauthorized')

# Retrieve email


def getUser(email):
    user = User.objects.filter(email=email).first()
    return user

# Validate Candidate


def isCandidateExist(email, examCohort_id):
    return Candidate.objects.filter(examCohort=examCohort_id, user__email=email).exists()

# Create MCQ question


def createMCQ(mcqData):
    serializer = MCQSerializer(data=mcqData)
    serializer.is_valid(raise_exception=True)
    serializer.save()
    return Response(data=serializer.data, status=status.HTTP_201_CREATED)

# Create MCQ Question's Choice


def createChoice(choiceData):
    serializer = MCQChoiceSerializer(data=choiceData)
    serializer.is_valid(raise_exception=True)
    serializer.save()
    return Response(status=status.HTTP_201_CREATED)

# Create Assessment


def createAssessment(assessmentData):
    print(assessmentData)
    serializer = AssessmentSerializer(data=assessmentData)
    serializer.is_valid(raise_exception=True)
    serializer.save()
    return Response(data=serializer.data, status=status.HTTP_201_CREATED)

# Create Miro-Viva Question


def createMicroViva(microVivaData):
    serializer = MicroVivaSerializer(data=microVivaData)
    serializer.is_valid(raise_exception=True)
    serializer.save()
    return Response(data=serializer.data, status=status.HTTP_201_CREATED)

# Validate Date and Time


def parseDateTime(dateTimeData):
    dateTimeFormat = "%Y-%m-%dT%H:%M:%S%z"
    dateTime = datetime.datetime.strptime(dateTimeData, dateTimeFormat)
    if dateTime is None:
        raise NotAcceptable(detail='Invalid DateTime format')
    convertedDateTime = dateTime.astimezone(tz=pytz.UTC)
    return convertedDateTime

# Validate Duration


def validateDuration(availableDateTime, dueDateTime, assessmentDuration):
    proposedDuration = dueDateTime - availableDateTime

    assessmentDurationTimeDelta = datetime.timedelta(
        seconds=assessmentDuration)

    maxDuration = datetime.timedelta(days=365)

    if proposedDuration >= maxDuration:
        raise NotAcceptable(detail='Total assessment duration is too long')

    dateTimeNow = datetime.datetime.now(tz=pytz.UTC)

    if ((availableDateTime - dateTimeNow) > maxDuration) or ((dueDateTime - dateTimeNow) > (maxDuration + maxDuration)):
        raise NotAcceptable(detail='Date and Time for the assessment too far')

    if availableDateTime < dateTimeNow:
        raise NotAcceptable(detail='Available date time is already past')

    if dueDateTime < dateTimeNow:
        raise NotAcceptable(detail='Due date time is already past')

    if (proposedDuration) < assessmentDurationTimeDelta:
        raise NotAcceptable(detail='Total assessment duration is too short')

    return True

# Count Total Duration for all MCQ question


def countTotalTimeForMCQ(mcqList):
    totalTimeInSecond = 0
    for mcq in mcqList:
        totalTimeInSecond += mcq.totalTime
    return totalTimeInSecond

# Count Total Duration for all Micro-viva question


def countTotalTimeForMicroViva(microVivaList):
    totalTimeInSecond = 0
    for microViva in microVivaList:
        totalTimeInSecond += microViva.totalTime
    return totalTimeInSecond


def isValidKeys(data, keys):
    try:
        for key in keys:
            dataVal = data[key]
        return True
    except KeyError:
        return False


def speechToText(audioFile):
    r = sr.Recognizer()
    with audioFile as source:
        audio = r.record(source)
    text = r.recognize_google(audio)
    return text


def validateQuestion(microViva):
    r = sr.Recognizer()
    # f = open(microViva.mvQuestionAudio.path, 'w')
    # fileobject = File(f)
    filename = 'out.wav'
    ffmpeg_tools.ffmpeg_extract_audio(microViva.mvQuestionAudio.path, filename)
    audioFile = sr.AudioFile(filename)
    with audioFile as source:
        audio = r.record(source)
    text = r.recognize_google(audio)
    # fileobject.close()
    # f.close()
    if text == '':
        raise NotAcceptable(detail='Question cannot be empty')
    microViva.mvQuestionText = text
    microViva.save()


def validateAnswer(microViva):
    r = sr.Recognizer()
    # f = open(microViva.evaluatorsAnswerAudio.path, 'w')
    # fileobject = File(f)
    filename = 'out.wav'
    ffmpeg_tools.ffmpeg_extract_audio(
        microViva.evaluatorsAnswerAudio.path, filename)
    audioFile = sr.AudioFile(filename)
    with audioFile as source:
        audio = r.record(source)
    text = r.recognize_google(audio)
    # fileobject.close()
    # f.close()
    if text == '':
        raise NotAcceptable(detail='Answer cannot be empty')
    if len(text.split(' ')) > 5:
        raise NotAcceptable(detail='Answer should be in 5 words')
    microViva.evaluatorsAnswerTEXT = text
    microViva.save()


def validateTime(microViva):
    filename1 = 'out1.wav'
    filename2 = 'out2.wav'

    ffmpeg_tools.ffmpeg_extract_audio(
        microViva.mvQuestionAudio.path, filename1)

    ffmpeg_tools.ffmpeg_extract_audio(
        microViva.evaluatorsAnswerAudio.path, filename2)

    audio_info1 = mutagen.File(filename1).info
    audio_info2 = mutagen.File(filename2).info

    if audio_info1.length + audio_info2.length > microViva.totalTime:
        raise NotAcceptable(detail='Time is too short to answer the question')
    if microViva.totalTime > 3600:
        raise NotAcceptable(detail='Time should not exceed 1 hour')


def publishAssessment(data):
    assessment = Assessment.objects.get(pk=data['assessmentId'])
    assessment.assessmentName = data['assessmentName']
    assessment.availableDateTime = data['availableDateTime']
    assessment.dueDateTime = data['dueDateTime']
    assessment.isPublished = True
    assessment.save()


def authorizeCandidate(request):
    payload = authorizeUser(request)
    candidate = Candidate.objects.filter(
        examCohort=request.data['id'], user=payload['id']).first()
    if candidate:
        payload['candidateId'] = candidate.pk
        return payload
    raise AuthenticationFailed(detail='Unauthorized')


def authorizeCandidateForExam(request):
    if not isValidKeys(request.data, ['assessmentId']):
        raise NotAcceptable()
    assessment = Assessment.objects.get(pk=request.data['assessmentId'])
    if not assessment.isPublished:
        raise NotAcceptable(detail='Assessment is not published yet')
    if assessment.availableDateTime > datetime.datetime.now(datetime.timezone.utc):
        raise NotAcceptable(detail='Assessment is not started yet')
    if assessment.dueDateTime < datetime.datetime.now(datetime.timezone.utc):
        raise NotAcceptable(detail='Due time passed')
    cohort = ExamCohort.objects.filter(
        assessment__id=request.data['assessmentId']).first()
    request.data['id'] = cohort.pk
    payload = authorizeCandidate(request)
    return payload


def submitMCQ(submissionData):
    submission = MCQSubmission.objects.get(pk=submissionData['id'])
    choice = MCQChoice.objects.get(pk=submissionData['choiceId'])
    mcq = MCQ.objects.get(pk=submissionData['questionId'])

    if not mcq or not submission or not choice:
        raise NotAcceptable(detail='Invalid Question')

    obtainedMarks = mcq.totalMarks if mcq.evaluatorsAnswer == choice.choice else 0

    submission.choice = choice
    submission.obtainedMarks = obtainedMarks
    submission.isSubmitted = True
    submission.save()


def submitMicroViva(submissionData):
    submission = MicroVivaSubmission.objects.get(pk=submissionData['id'])
    mv = MicroViva.objects.get(pk=submissionData['questionId'])

    if not mv or not submission:
        raise NotAcceptable(detail='Invalid Question')

    submission.submittedAnswerAudio = submissionData['mvAnswer']
    submission.save()

    filename = 'out.wav'
    ffmpeg_tools.ffmpeg_extract_audio(
        submission.submittedAnswerAudio.path, filename)

    submission.submittedAnswerText = speechToText(
        sr.AudioFile(filename))
    submission.save()

    submission.obtainedMarks = mv.totalMarks if is_correct_answer(
        mv.mvQuestionText, mv.evaluatorsAnswerTEXT, submission.submittedAnswerText) else 0
    submission.isSubmitted = True
    submission.save()


# Regiser View of Google User


class RegisterView(APIView):
    def post(self, request):
        token = request.data['token']
        try:
            # Specify the CLIENT_ID of the app that accesses the backend:
            idinfo = id_token.verify_oauth2_token(
                token, requests.Request(), CLIENT_ID, clock_skew_in_seconds=10)

            # Or, if multiple clients access the backend server:
            # idinfo = id_token.verify_oauth2_token(token, requests.Request())
            # if idinfo['aud'] not in [CLIENT_ID_1, CLIENT_ID_2, CLIENT_ID_3]:
            #     raise ValueError('Could not verify audience.')

            # If auth request is from a G Suite domain:
            # if idinfo['hd'] != GSUITE_DOMAIN_NAME:
            #     raise ValueError('Wrong hosted domain.')

            # ID token is valid. Get the user's Google Account ID from the decoded token.
            user = User.objects.filter(sub=idinfo['sub']).first()
            if user is not None:
                raise NotAcceptable(detail='User already exists')
            data = {'name': idinfo['name'],
                    'email': idinfo['email'], 'sub': idinfo['sub']}
            serializer = UserSerializer(data=data)
            serializer.is_valid(raise_exception=True)
            serializer.save()
            return Response(status=status.HTTP_201_CREATED)
        except ValueError:
            raise AuthenticationFailed('Invalid credentials')

# Login View


class LoginView(APIView):
    def post(self, request):
        if not isValidKeys(request.data, ['token']):
            raise NotAcceptable()
        token = request.data['token']
        for i in range(10):
            try:
                # Specify the CLIENT_ID of the app that accesses the backend:
                idinfo = id_token.verify_oauth2_token(
                    token, requests.Request(), CLIENT_ID, clock_skew_in_seconds=10)

                # Or, if multiple clients access the backend server:
                # idinfo = id_token.verify_oauth2_token(token, requests.Request())
                # if idinfo['aud'] not in [CLIENT_ID_1, CLIENT_ID_2, CLIENT_ID_3]:
                #     raise ValueError('Could not verify audience.')

                # If auth request is from a G Suite domain:
                # if idinfo['hd'] != GSUITE_DOMAIN_NAME:
                #     raise ValueError('Wrong hosted domain.')

                # ID token is valid. Get the user's Google Account ID from the decoded token.
                sub = idinfo['sub']
                user = User.objects.filter(sub=sub).first()
                if user is None:
                    registerApi = RegisterView()
                    registerApi.post(request=request)
                    user = User.objects.filter(sub=sub).first()
                payload = {
                    'id': user.id,
                    'iat': datetime.datetime.now(tz=pytz.UTC)
                }

                token = jwt.encode(payload, SECRET_KEY, algorithm='HS256')

                response = Response()
                response.set_cookie(key='jwt', value=token,
                                    httponly=True, samesite=None)
                response.set_cookie(
                    key='isLoggedIn', value='Yes', samesite=None)
                response.status_code = status.HTTP_200_OK
                return response

            except ValueError:
                # Invalid token
                if i > 8:
                    raise AuthenticationFailed(detail='Invalid credentials')

# User view


class UserView(APIView):
    def get(self, request):
        payload = authorizeUser(request)
        user = User.objects.filter(id=payload['id']).first()
        serializer = UserSerializer(user)
        return Response(serializer.data)

# Logout View


class LogoutView(APIView):
    def post(self, request):
        response = Response()
        response.delete_cookie('jwt')
        response.delete_cookie('isLoggedIn')
        response.data = {
            'message': 'success'
        }
        response.status_code = status.HTTP_200_OK
        return response

# Evaluator's Dashboard view


class EvaluatorDashboardView(APIView):
    def get(self, request):
        payload = authorizeUser(request)
        user = User.objects.filter(id=payload['id']).first()
        queryset = ExamCohort.objects.filter(evaluator__user=user.id)
        return Response(data=queryset.values(), status=status.HTTP_200_OK)

# Candidate's Dashboard View


class CandidateDashboardView(APIView):
    def get(self, request):
        payload = authorizeUser(request)
        user = User.objects.filter(id=payload['id']).first()
        queryset = ExamCohort.objects.filter(candidate__user=user.id)
        return Response(data=queryset.values(), status=status.HTTP_200_OK)

# Create Exam Cohort View


class CreateExamCohortView(APIView):
    def post(self, request):
        payload = authorizeUser(request)
        cohortSerializer = ExamCohortSerializer(data=request.data)
        cohortSerializer.is_valid(raise_exception=True)
        cohortSerializer.save()
        data = {
            'user': payload['id'],
            'examCohort': cohortSerializer.data['id']
        }
        evaluatorSerializer = EvaluatorSerializer(data=data)
        evaluatorSerializer.is_valid(raise_exception=True)
        evaluatorSerializer.save()
        return Response(status=status.HTTP_201_CREATED)

# Add Candidate view


class AddCandidateView(APIView):
    def post(self, request):
        if not isValidKeys(request.data, ['email', 'id']):
            raise NotAcceptable()
        payload = authorizeEvaluator(request)
        print(request.data)
        email = request.data['email']
        examCohort_id = request.data['id']
        if email is None or examCohort_id is None:
            raise ParseError()
        user = getUser(request.data['email'])
        if user is not None:
            if isCandidateExist(email, examCohort_id):
                raise NotAcceptableE(detail='Candidate already exists')
            if payload['id'] == user.id:
                raise NotAcceptable(detail='Evaluator cannot be a Candidate')
            data = {
                'examCohort': examCohort_id,
                'user': user.id
            }
            serializer = CandidateSerializer(data=data)
            serializer.is_valid(raise_exception=True)
            serializer.save()
            return Response(data={'message : successfully added'}, status=status.HTTP_202_ACCEPTED)
        raise NotFound(detail={'message': 'User not found'})

# Create Assessment View


class CreateAssessmentView(APIView):
    def post(self, request):
        if not isValidKeys(request.data, ['name', 'id', 'availableDateTime', 'dueDateTime']):
            raise NotAcceptable()
        payload = authorizeEvaluator(request)
        validateDuration(
            parseDateTime(request.data['availableDateTime']), parseDateTime(request.data['dueDateTime']), 0)
        data = {
            'assessmentName': request.data['name'],
            'examCohort': request.data['id'],
            'availableDateTime': parseDateTime(request.data['availableDateTime']),
            'dueDateTime': parseDateTime(request.data['dueDateTime']),
        }
        assessmentResponse = createAssessment(data)

        return Response(data=assessmentResponse.data, status=status.HTTP_201_CREATED)


class CreateMCQView(APIView):
    def post(self, request):
        if not isValidKeys(request.data, ['assessmentId', 'mcqQuestion', 'choiceList', 'time', 'mcqMarks', 'correctAnswer']):
            raise NotAcceptable()
        cohort = ExamCohort.objects.filter(
            assessment__id=request.data['assessmentId']).first()
        request.data['id'] = cohort.pk
        authorizeEvaluator(request)
        if len(request.data['choiceList']) < 2:
            raise NotAcceptable(detail='There should be at least 2 choices')
        if request.data['correctAnswer'] not in request.data['choiceList']:
            raise NotAcceptable(detail='Invalid correct answer')
        mcqData = {
            'questionGroup': request.data['questionGroup'],
            'assessment': request.data['assessmentId'],
            'MCQQuestion': request.data['mcqQuestion'],
            'evaluatorsAnswer': request.data['correctAnswer'],
            'totalMarks': request.data['mcqMarks'],
            'totalTime': request.data['time']
        }
        mcqResponse = createMCQ(mcqData)
        choiceList = request.data['choiceList']
        choiceList.append('5a72c3f196603e351736f4ffdd5985cc')
        for choice in choiceList:
            choiceData = {
                'mcq': mcqResponse.data['id'],
                'choice': choice
            }
            createChoice(choiceData)
        return Response(data='MCQ Created', status=status.HTTP_201_CREATED)


class CreateMicroVivaView(APIView):
    def post(self, request):
        print(request.data)
        with transaction.atomic():
            if not isValidKeys(request.data, ['assessmentId', 'mvQuestionAudio', 'evaluatorsAnswerAudio', 'time', 'mvMarks']):
                raise NotAcceptable()
            cohort = ExamCohort.objects.filter(
                assessment__id=request.data['assessmentId']).first()
            request.data['id'] = cohort.pk
            authorizeEvaluator(request)
            data = {
                'questionGroup': request.data['questionGroup'],
                'assessment': request.data['assessmentId'],
                'mvQuestionAudio': request.data['mvQuestionAudio'],
                'evaluatorsAnswerAudio': request.data['evaluatorsAnswerAudio'],
                'totalTime': request.data['time'],
                'totalMarks': request.data['mvMarks']
            }
            microVivaResponse = createMicroViva(data)
            microViva = MicroViva.objects.get(pk=microVivaResponse.data['id'])
            validateQuestion(microViva)
            validateAnswer(microViva)
            validateTime(microViva)
            return Response(data='MicroViva created', status=status.HTTP_201_CREATED)


class PublishAssessmentView(APIView):
    def post(self, request):
        if not isValidKeys(request.data, ['name', 'id', 'availableDateTime', 'dueDateTime', 'assessmentId']):
            raise NotAcceptable(detail="Invalid Request Header")
        payload = authorizeEvaluator(request)
        data = {
            'assessmentId': request.data['assessmentId'],
            'assessmentName': request.data['name'],
            'examCohort': request.data['id'],
            'availableDateTime': parseDateTime(request.data['availableDateTime']),
            'dueDateTime': parseDateTime(request.data['dueDateTime']),
        }
        mcqQueryset = MCQ.objects.filter(
            assessment__id=request.data['assessmentId'])
        mvQueryset = MicroViva.objects.filter(
            assessment__id=request.data['assessmentId'])
        assessmentDuration = countTotalTimeForMCQ(
            mcqQueryset) + countTotalTimeForMicroViva(mvQueryset)
        print(assessmentDuration)
        if assessmentDuration == 0:
            raise NotAcceptable(detail='Add some question before publishing')
        validateDuration(data['availableDateTime'],
                         data['dueDateTime'], assessmentDuration)
        assessmentResponse = publishAssessment(data)

        return Response(data='Assessment Published', status=status.HTTP_200_OK)


class AssessmentView(APIView):
    def post(self, request):
        payload = authorizeEvaluator(request)
        examCohort = request.data['id']
        queryset = Assessment.objects.filter(
            examCohort__id=examCohort)
        return Response(data=queryset.values(), status=status.HTTP_200_OK)


class AssessmentCandidateView(APIView):
    def post(self, request):
        payload = authorizeCandidate(request)
        examCohort = request.data['id']
        queryset = Assessment.objects.filter(
            examCohort__id=examCohort, isPublished=True)
        return Response(data=queryset.values(), status=status.HTTP_200_OK)


class SingleAssessmentView(APIView):
    def post(self, request):
        if not isValidKeys(request.data, ['id', 'assessmentId']):
            raise NotAcceptable()
        payload = authorizeEvaluator(request)
        examCohort = request.data['id']
        queryset = Assessment.objects.filter(
            examCohort__id=examCohort, pk=request.data['assessmentId'])
        if queryset.count() == 0:
            raise AuthenticationFailed(detail='Unauthorized')
        return Response(data=queryset.values(), status=status.HTTP_200_OK)


class SingleAssessmentCandidateView(APIView):
    def post(self, request):
        if not isValidKeys(request.data, ['id', 'assessmentId']):
            raise NotAcceptable()
        payload = authorizeCandidate(request)
        examCohort = request.data['id']
        queryset = Assessment.objects.filter(
            examCohort__id=examCohort, pk=request.data['assessmentId'])
        if queryset.count() == 0:
            raise AuthenticationFailed(detail='Unauthorized')
        return Response(data=queryset.values(), status=status.HTTP_200_OK)


class SearchUserView(APIView):
    def post(self, request):
        payload = authorizeEvaluator(request)
        evaluator = payload['id']
        examCohort = request.data['id']
        pattern = request.data['pattern']
        queryset = User.objects.filter(email__contains=pattern)
        queryset2 = Candidate.objects.filter(examCohort__id=examCohort)
        finalQueryset = queryset.filter(
            ~Q(id__in=Subquery(queryset2.values('user__id'))), ~Q(id=evaluator))

        return Response(data=finalQueryset[:5].values(), status=status.HTTP_200_OK)


class CandidateView(APIView):
    def post(self, request):
        payload = authorizeEvaluator(request)
        examCohort = request.data['id']
        queryset = User.objects.filter(candidate__examCohort=examCohort)

        return Response(data=queryset.values('id', 'name', 'email'), status=status.HTTP_200_OK)


class CreateQuestionGroupView(APIView):
    def post(self, request):
        print(request.data)
        payload = authorizeEvaluator(request)
        data = {
            'name': request.data['name'],
            'assessment': request.data['assessment']
        }
        serializer = QuestionGroupSerializer(data=data)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(data=serializer.data, status=status.HTTP_201_CREATED)


class GetQuestionView(APIView):
    def post(self, request):
        payload = authorizeCandidateForExam(request)

        unsubmittedMCQ = MCQ.objects.filter(mcqsubmission__candidate__id=payload['candidateId'], mcqsubmission__submissionEndTime__gt=datetime.datetime.now(
            datetime.timezone.utc), mcqsubmission__isSubmitted=0, assessment__id=request.data['assessmentId']).values('id', 'MCQQuestion')
        if unsubmittedMCQ:
            data = unsubmittedMCQ[0]
            return Response(data=data, status=status.HTTP_200_OK)

        unsubmittedMV = MicroViva.objects.filter(microvivasubmission__candidate__id=payload['candidateId'], microvivasubmission__submissionEndTime__gt=datetime.datetime.now(
            datetime.timezone.utc), microvivasubmission__isSubmitted=0, assessment__id=request.data['assessmentId']).values('id')

        if unsubmittedMV:
            data = unsubmittedMV[0]
            print(data)
            return Response(data=data, status=status.HTTP_200_OK)

        last_submission = MCQSubmission.objects.filter(
            assessment__id=request.data['assessmentId'], mcqsubmission__candidate__id=payload['candidateId']).latest('submittedAt')
        if last_submission:
            last_submitted_mcq = MCQ.objects.filter(pk=last_submission.mcq_id)
            left_in_group = MCQ.objects.filter(questionGroup=last_submitted_mcq.questionGroup,
                                               created_at__gt=last_submitted_mcq.createdAt).order_by('createdAt').first()
            if left_in_group:
                return Response(data=left_in_group, status=status.HTTP_200_OK)
            else:
                data = {
                    'user': payload['candidateId'],
                    'group': last_submitted_mcq.questionGroup
                }
                user_and_group = UserAndGroupSerializer(data=data)
                user_and_group.is_valid(raise_exception=True)
                user_and_group.save()

        group_list = QuestionGroup.objects.filter(
            assessment__id=request.data['assessmentId'], isDone=False)

        group_id = group_list[randint(0, group_list.count()-1)]['id']
        group_submission = UserAndGroup.objects.filter(
            user__id=payload['candidateId'], group__id=group_id)
        while (len(group_submission) != 0):
            group_id = group_list[randint(0, group_list.count()-1)]['id']
            group_submission = UserAndGroup.objects.filter(
                user__id=payload['candidateId'], group__id=group_id).values()
            temp_group_submission = UserAndGroup.objects.filter(
                user__id=payload['candidateId'])

        mcqList = list(MCQ.objects.filter(
            ~Q(mcqsubmission__candidate__id=payload['candidateId']), assessment__id=request.data['assessmentId']).values('id', 'MCQQuestion'))
        mvList = list(MicroViva.objects.filter(
            ~Q(microvivasubmission__candidate__id=payload['candidateId']), assessment__id=request.data['assessmentId']).values('id'))
        questionList = mcqList+mvList
        data = {}
        response = Response(status=status.HTTP_200_OK)
        # print(MicroVivaSubmission.objects.all().values())
        if not questionList:
            data['isFinished'] = 'Yes'
            response.data = data
        else:
            print(questionList)
            data = questionList[randint(0, len(questionList)-1)]
            response.data = data

        return response


class GetMCQChoiceView(APIView):
    def post(self, request):
        payload = authorizeCandidate(request)
        if not isValidKeys(request.data, ['questionId']):
            raise NotAcceptable()

        choices = list(MCQChoice.objects.filter(
            mcq__id=request.data['questionId']).values())
        random.shuffle(choices)
        data = {
            'choiceList': choices
        }
        return Response(data=data, status=status.HTTP_200_OK)


class GetMicroVivaQuestionView(APIView):
    def post(self, request):
        payload = authorizeCandidate(request)
        if not isValidKeys(request.data, ['questionId']):
            raise NotAcceptable()

        mv = MicroViva.objects.get(pk=request.data['questionId'])
        print(mv.mvQuestionAudio.path)
        response = FileResponse(open(mv.mvQuestionAudio.path, 'rb'))
        response.filename = mv.mvQuestionAudio.name
        response.status_code = status.HTTP_200_OK

        return response


class GetAudioFileView(APIView):
    def post(self, request):
        payload = authorizeEvaluator(request)
        response = FileResponse(open(request.data['path'], 'rb'))
        response.status_code = status.HTTP_200_OK
        return response


class StartMCQSubmissionView(APIView):
    def post(self, request):
        payload = authorizeCandidate(request)
        if not isValidKeys(request.data, ['questionId']):
            raise NotAcceptable()

        queryset = MCQSubmission.objects.filter(
            mcq__id=request.data['questionId'], candidate__id=payload['candidateId'])

        submission = ""
        if queryset.exists():
            submission = queryset.first()
        if (submission):
            availableTime = submission.submissionEndTime - \
                datetime.datetime.now(datetime.timezone.utc)
            data = {
                'remainingTime': availableTime.seconds
            }
            return Response(data=data, status=status.HTTP_200_OK)

        mcq = MCQ.objects.get(pk=request.data['questionId'])
        submissionData = {
            'mcq': mcq.id,
            'candidate': payload['candidateId'],
            'submissionStartTime': datetime.datetime.now(datetime.timezone.utc),
            'submissionEndTime': datetime.datetime.now(datetime.timezone.utc) + datetime.timedelta(seconds=mcq.totalTime)
        }
        serializer = MCQSubmissionSerializer(data=submissionData)
        serializer.is_valid(raise_exception=True)
        serializer.save()

        return Response(data={'remainingTime': mcq.totalTime}, status=status.HTTP_200_OK)


class StartMVSubmissionView(APIView):
    def post(self, request):
        payload = authorizeCandidate(request)
        if not isValidKeys(request.data, ['questionId']):
            raise NotAcceptable()

        queryset = MicroVivaSubmission.objects.filter(
            mv__id=request.data['questionId'], candidate__id=payload['candidateId'])

        submission = ""
        if queryset.exists():
            submission = queryset.first()
        print(submission)
        if (submission):
            availableTime = submission.submissionEndTime - \
                datetime.datetime.now(datetime.timezone.utc)
            data = {
                'remainingTime': availableTime.seconds
            }
            return Response(data=data, status=status.HTTP_200_OK)

        mv = MicroViva.objects.get(pk=request.data['questionId'])
        submissionData = {
            'mv': mv.id,
            'candidate': payload['candidateId'],
            'submissionStartTime': datetime.datetime.now(datetime.timezone.utc),
            'submissionEndTime': datetime.datetime.now(datetime.timezone.utc) + datetime.timedelta(seconds=mv.totalTime)
        }
        serializer = MicroVivaSubmissionSerializer(data=submissionData)
        serializer.is_valid(raise_exception=True)
        serializer.save()

        return Response(data={'remainingTime': mv.totalTime}, status=status.HTTP_200_OK)


class SubmitMCQSubmissionView(APIView):
    def post(self, request):
        payload = authorizeCandidate(request)
        if not isValidKeys(request.data, ['questionId', 'choiceId']):
            raise NotAcceptable()
        submission = MCQSubmission.objects.get(
            mcq__id=request.data['questionId'], candidate__id=payload['candidateId'])
        if not submission:
            raise NotAcceptable(detail='Submission is not initiated')
        if submission.submissionEndTime + datetime.timedelta(seconds=5) < datetime.datetime.now(datetime.timezone.utc):
            raise NotAcceptable(detail='Submission time expired')
        assessment = Assessment.objects.get(mcq__id=request.data['questionId'])
        if assessment.dueDateTime + datetime.timedelta(seconds=5) < datetime.datetime.now(datetime.timezone.utc):
            raise NotAcceptable(detail='Due Date Time is passed')
        submissionData = {
            'id': submission.id,
            'choiceId': request.data['choiceId'],
            'questionId': request.data['questionId']
        }
        submitMCQ(submissionData)
        return Response(status=status.HTTP_202_ACCEPTED)


class SubmitMicroMivaSubmissionView(APIView):
    def post(self, request):
        payload = authorizeCandidate(request)
        if not isValidKeys(request.data, ['questionId', 'mvAnswer']):
            raise NotAcceptable()
        submission = MicroVivaSubmission.objects.get(
            mv__id=request.data['questionId'], candidate__id=payload['candidateId'])
        if not submission:
            raise NotAcceptable(detail='Submission is not initiated')
        if submission.submissionEndTime + datetime.timedelta(seconds=5) < datetime.datetime.now(datetime.timezone.utc):
            raise NotAcceptable(detail='Submission time expired')

        assessment = Assessment.objects.get(
            microviva__id=request.data['questionId'])
        if assessment.dueDateTime + datetime.timedelta(seconds=5) < datetime.datetime.now(datetime.timezone.utc):
            raise NotAcceptable(detail='Due Date Time is passed')

        submissionData = {
            'id': submission.id,
            'questionId': request.data['questionId'],
            'mvAnswer': request.data['mvAnswer']
        }

        submitMicroViva(submissionData)
        return Response(status=status.HTTP_202_ACCEPTED)


class GetMarksView(APIView):
    def post(self, request):
        payload = authorizeCandidate(request)

        mcqMarks = MCQSubmission.objects.filter(mcq__assessment__id=request.data['assessmentId'], candidate__id=payload['candidateId']).aggregate(
            Sum('obtainedMarks'))['obtainedMarks__sum']
        mvMarks = MicroVivaSubmission.objects.filter(
            mv__assessment__id=request.data['assessmentId'], candidate__id=payload['candidateId']).aggregate(Sum('obtainedMarks'))['obtainedMarks__sum']

        mcqTotalMarks = MCQ.objects.filter(assessment__id=request.data['assessmentId']).aggregate(
            Sum('totalMarks'))['totalMarks__sum']
        mvTotalMarks = MicroViva.objects.filter(assessment__id=request.data['assessmentId']).aggregate(
            Sum('totalMarks'))['totalMarks__sum']

        obtainedMarks = 0
        obtainedMarks += (mcqMarks if mcqMarks else 0)
        obtainedMarks += (mvMarks if mvMarks else 0)

        totalMarks = 0
        totalMarks += (mcqTotalMarks if mcqTotalMarks else 0)
        totalMarks += (mvTotalMarks if mvTotalMarks else 0)

        print(mcqMarks, mvMarks, mcqTotalMarks, mvTotalMarks)

        return Response(data={'obtainedMarks': obtainedMarks, 'totalMarks': totalMarks}, status=status.HTTP_200_OK)


class GetSubmissionView(APIView):
    def post(self, request):
        print(request.data)
        payload = authorizeEvaluator(request)
        mcqSubmissionList = MCQSubmission.objects.filter(
            mcq__assessment__id=request.data['assessmentId'], candidate__user__id=request.data['candidateId']).values()
        mvSubmissionList = MicroVivaSubmission.objects.filter(
            mv__assessment__id=request.data['assessmentId'], candidate__user__id=request.data['candidateId']).values()
        totalMarks = 0
        obtainedMarks = 0
        for i in range(len(mcqSubmissionList)):
            obtainedMarks += mcqSubmissionList[i]['obtainedMarks']
            mcqSubmissionList[i]['mcqQuestion'] = MCQ.objects.filter(
                id=mcqSubmissionList[i]['mcq_id']).values()
            totalMarks += mcqSubmissionList[i]['mcqQuestion'][0]['totalMarks']
            mcqSubmissionList[i]['mcqChoiceList'] = MCQChoice.objects.filter(
                mcq__id=mcqSubmissionList[i]['mcq_id']).values()
            for mcqChoice in mcqSubmissionList[i]['mcqChoiceList']:
                if mcqChoice['id'] == mcqSubmissionList[i]['choice_id']:
                    mcqSubmissionList[i]['candidatesAnswer'] = mcqChoice['choice']
        for i in range(len(mvSubmissionList)):
            obtainedMarks += mvSubmissionList[i]['obtainedMarks']
            mvSubmissionList[i]['mvQuestion'] = MicroViva.objects.filter(
                id=mvSubmissionList[i]['mv_id']).values()
            totalMarks += mvSubmissionList[i]['mvQuestion'][0]['totalMarks']

        data = {
            'mcqSubmissionList': mcqSubmissionList,
            'mvSubmissionList': mvSubmissionList,
            'obtainedMarks': obtainedMarks,
            'totalMarks': totalMarks
        }
        return Response(data=data, status=status.HTTP_200_OK)


class GetSpreadsheetView(APIView):
    def post(self, request):
        if not isValidKeys(request.data, ['id']):
            raise NotAcceptable()
        payload = authorizeEvaluator(request)

        examCohort = ExamCohort.objects.get(pk=request.data['id'])
        candidateList = list(User.objects.filter(
            candidate__examCohort__id=examCohort.pk).values())
        assessmentList = list(Assessment.objects.filter(
            examCohort__id=examCohort.pk, isPublished=True, dueDateTime__lt=datetime.datetime.now(datetime.timezone.utc)).values())

        wb = Workbook()
        ws = wb.active
        ws.cell(1, 1, "ID")
        ws.cell(1, 2, "Name")
        row = 1
        col = 2
        for assessment in assessmentList:
            col += 1
            ws.cell(row, col, assessment['assessmentName'])
        for candidate in candidateList:
            row += 1
            col = 2
            ws.cell(row, 1, row)
            ws.cell(row, 2, candidate['name'])
            for assessment in assessmentList:
                col += 1
                mcqMarks = MCQSubmission.objects.filter(mcq__assessment__id=assessment['id'], candidate__id=candidate['id']).aggregate(
                    Sum('obtainedMarks'))['obtainedMarks__sum']
                mvMarks = MicroVivaSubmission.objects.filter(
                    mv__assessment__id=assessment['id'], candidate__id=candidate['id']).aggregate(Sum('obtainedMarks'))['obtainedMarks__sum']
                ws.cell(row, col, mcqMarks+mvMarks)
        filename = f'./Marksheet/Marksheet-{examCohort.id}.xlsx'
        wb.save(filename)
        response = FileResponse(open(filename, 'rb'))
        response.filename = filename
        response.status_code = status.HTTP_200_OK
        # os.remove(filename)
        return response


class GetQuestionListView(APIView):
    def post(self, request):
        if not isValidKeys(request.data, ['id', 'assessmentId']):
            raise NotAcceptable()

        payload = authorizeEvaluator(request)
        if not Assessment.objects.filter(examCohort__id=request.data['id']).exists():
            raise AuthenticationFailed('Unauthorized')
        mcqList = list(MCQ.objects.filter(
            assessment__id=request.data['assessmentId']).values('id', 'MCQQuestion', 'questionGroup'))
        mvList = list(MicroViva.objects.filter(
            assessment__id=request.data['assessmentId']).values('id', 'questionGroup'))

        return Response(data={"mcqList": mcqList, "mvList": mvList})


class DeleteMCQView(APIView):
    def post(self, request):
        if not isValidKeys(request.data, ['id', 'assessmentId', 'questionId']):
            raise NotAcceptable()

        payload = authorizeEvaluator(request)
        if not Assessment.objects.filter(examCohort__id=request.data['id']).exists():
            raise AuthenticationFailed('Unauthorized')

        mcq = MCQ.objects.get(pk=request.data['questionId'])
        mcq.delete()

        return Response(data={"message": "success"}, status=status.HTTP_200_OK)


class DeleteMVView(APIView):
    def post(self, request):
        if not isValidKeys(request.data, ['id', 'assessmentId', 'questionId']):
            raise NotAcceptable()

        payload = authorizeEvaluator(request)
        if not Assessment.objects.filter(examCohort__id=request.data['id']).exists():
            raise AuthenticationFailed('Unauthorized')

        mv = MicroViva.objects.get(pk=request.data['questionId'])
        mv.delete()

        return Response(data={"message": "success"}, status=status.HTTP_200_OK)


class UpdateMCQMarksView(APIView):
    def post(self, request):
        payload = authorizeEvaluator(request)
        mcqSubmission = MCQSubmission.objects.get(
            id=request.data['mcqSubmissionId'])
        mcqSubmission.obtainedMarks = (
            request.data['totalMarks'] if mcqSubmission.obtainedMarks == 0 else 0)
        mcqSubmission.save()

        return Response(status=status.HTTP_200_OK)


class UpdateMvMarksView(APIView):
    def post(self, request):
        payload = authorizeEvaluator(request)
        mvSubmission = MicroVivaSubmission.objects.get(
            id=request.data['mvSubmissionId'])
        mvSubmission.obtainedMarks = (
            request.data['totalMarks'] if mvSubmission.obtainedMarks == 0 else 0)
        mvSubmission.save()

        return Response(status=status.HTTP_200_OK)
